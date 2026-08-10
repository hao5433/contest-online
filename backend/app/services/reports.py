"""Builds the Excel (openpyxl) and PDF (reportlab) exam reports, sharing the
same tabular data so both formats stay in sync."""
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.models.exam_attempt import ExamAttempt

REPORT_HEADERS = ["Họ tên", "Email", "Điểm", "Thời gian nộp", "Thời gian làm (phút)", "Số lần vi phạm"]


def _report_rows(attempts: list[ExamAttempt]) -> list[list]:
    rows = []
    for attempt in attempts:
        duration_taken: float | str = "-"
        if attempt.submitted_at:
            delta = attempt.submitted_at - attempt.started_at
            duration_taken = round(delta.total_seconds() / 60, 1)
        rows.append(
            [
                attempt.student.full_name,
                attempt.student.email,
                attempt.score if attempt.score is not None else "-",
                attempt.submitted_at.strftime("%Y-%m-%d %H:%M:%S") if attempt.submitted_at else "-",
                duration_taken,
                attempt.violation_count,
            ]
        )
    return rows


def build_excel_report(exam_title: str, attempts: list[ExamAttempt]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.append([f"Báo cáo kết quả thi: {exam_title}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(REPORT_HEADERS)
    for cell in ws[3]:
        cell.font = Font(bold=True)
    for row in _report_rows(attempts):
        ws.append(row)

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = max(12, min(40, max_length + 2))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def build_pdf_report(exam_title: str, attempts: list[ExamAttempt]) -> io.BytesIO:
    """Note: uses reportlab's built-in Helvetica font, which only covers
    WinAnsi/Latin-1 glyphs. Vietnamese diacritics outside that range may not
    render correctly; embedding a Unicode TTF (e.g. Noto Sans) via
    reportlab.pdfbase.ttfonts is the production fix, left as a follow-up."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()

    elements = [
        Paragraph(f"Báo cáo kết quả thi: {exam_title}", styles["Title"]),
        Spacer(1, 12),
    ]
    data = [REPORT_HEADERS, *_report_rows(attempts)]
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2c3e50")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer
